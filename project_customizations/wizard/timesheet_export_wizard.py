import base64
import io
import re
import zipfile
from collections import defaultdict
from datetime import date, timedelta

from odoo import Command, api, fields, models
from odoo.exceptions import UserError

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError:
    Workbook = None


class TimesheetExportWizard(models.TransientModel):
    _name = 'timesheet.export.wizard'
    _description = 'Export Timesheet per Cliente'

    state = fields.Selection([
        ('draft', 'Bozza'),
        ('done', 'Completato'),
    ], default='draft')

    project_stage_ids = fields.Many2many(
        'project.project.stage',
        string='Fasi progetto',
        domain=[('is_to_invoice', '=', True)],
    )
    task_stage_name = fields.Char(
        string='Nome fase lavoro',
        help='I task in fasi con questo nome saranno inclusi nell\'export. '
             'Cerca per corrispondenza esatta (case insensitive).',
    )
    date_from = fields.Date(string='Data inizio')
    date_to = fields.Date(string='Data fine')
    line_count = fields.Integer(
        string='Righe da esportare',
        compute='_compute_line_count',
    )
    exported_count = fields.Integer(
        string='Righe esportate',
        readonly=True,
    )
    zip_file = fields.Binary(
        string='File ZIP',
        readonly=True,
        attachment=False,
    )
    zip_filename = fields.Char(string='Nome file')

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        if 'project_stage_ids' in fields_list:
            stages = self.env['project.project.stage'].search([
                ('is_to_invoice', '=', True),
            ])
            res['project_stage_ids'] = [Command.set(stages.ids)]
        if 'task_stage_name' in fields_list:
            res['task_stage_name'] = 'Da fatturare'
        # Default: primo e ultimo giorno del mese precedente
        today = date.today()
        first_of_current = today.replace(day=1)
        last_of_prev = first_of_current - timedelta(days=1)
        first_of_prev = last_of_prev.replace(day=1)
        if 'date_from' in fields_list:
            res['date_from'] = first_of_prev
        if 'date_to' in fields_list:
            res['date_to'] = last_of_prev
        return res

    @api.depends('project_stage_ids', 'task_stage_name', 'date_from', 'date_to')
    def _compute_line_count(self):
        for wizard in self:
            wizard.line_count = len(wizard._get_exportable_lines())

    def _get_task_stage_ids(self):
        """Trova tutte le fasi task che matchano il nome configurato."""
        if not self.task_stage_name:
            return []
        stages = self.env['project.task.type'].search([
            ('name', 'ilike', self.task_stage_name.strip()),
        ])
        return stages.ids

    def _get_exportable_lines(self):
        self.ensure_one()
        project_stage_ids = self.project_stage_ids.ids
        task_stage_ids = self._get_task_stage_ids()

        if not project_stage_ids and not task_stage_ids:
            return self.env['account.analytic.line']

        domain = [
            ('project_id', '!=', False),
            ('task_id', '!=', False),
        ]
        if self.date_from:
            domain.append(('date', '>=', self.date_from))
        if self.date_to:
            domain.append(('date', '<=', self.date_to))
        if project_stage_ids and task_stage_ids:
            stage_domain = [
                '|',
                ('project_id.stage_id', 'in', project_stage_ids),
                ('task_id.stage_id', 'in', task_stage_ids),
            ]
        elif project_stage_ids:
            stage_domain = [
                ('project_id.stage_id', 'in', project_stage_ids),
            ]
        else:
            stage_domain = [
                ('task_id.stage_id', 'in', task_stage_ids),
            ]
        return self.env['account.analytic.line'].search(
            domain + stage_domain, order='date asc',
        )

    def action_export(self):
        self.ensure_one()
        if Workbook is None:
            raise UserError(
                'La libreria openpyxl non è installata. '
                'Contattare l\'amministratore di sistema.'
            )
        lines = self._get_exportable_lines()
        if not lines:
            raise UserError(
                'Nessuna riga timesheet da esportare. '
                'Verificare che le fasi siano configurate come "Da Fatturare" '
                'e che ci siano fogli ore non ancora esportati.'
            )

        # Raggruppa per commercial partner
        grouped = defaultdict(lambda: self.env['account.analytic.line'])
        for line in lines:
            partner = line.partner_id.commercial_partner_id or line.partner_id
            grouped[partner] |= line

        # Genera ZIP
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for partner, partner_lines in grouped.items():
                excel_bytes = self._generate_excel(partner, partner_lines)
                filename = self._sanitize_filename(partner.name) + '.xlsx'
                zf.writestr(filename, excel_bytes)

        # Salva ZIP nel wizard
        zip_buffer.seek(0)
        timestamp = fields.Datetime.now().strftime('%Y%m%d_%H%M%S')
        self.write({
            'state': 'done',
            'exported_count': len(lines),
            'zip_file': base64.b64encode(zip_buffer.getvalue()),
            'zip_filename': f'export_timesheet_{timestamp}.zip',
        })

        return self._reopen_wizard()

    def _generate_excel(self, partner, lines):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Timesheet'

        headers = [
            'Data', 'Dipendente', 'Quantità', 'Progetto',
            'Lavoro', 'Descrizione', 'Voce ordine di vendita',
        ]
        bold = Font(bold=True)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = bold

        for row, line in enumerate(lines.sorted('date'), 2):
            ws.cell(row=row, column=1, value=line.date)
            ws.cell(row=row, column=1).number_format = 'DD/MM/YYYY'
            ws.cell(row=row, column=2, value=line.employee_id.name or '')
            ws.cell(row=row, column=3, value=line.unit_amount)
            ws.cell(row=row, column=4, value=line.project_id.name or '')
            ws.cell(row=row, column=5, value=line.task_id.name or '')
            ws.cell(row=row, column=6, value=line.name or '')
            ws.cell(row=row, column=7, value=line.so_line.name if line.so_line else '')

        # Larghezza colonne
        col_widths = [12, 20, 10, 25, 25, 40, 30]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def _sanitize_filename(name):
        if not name:
            return 'senza_nome'
        name = re.sub(r'[<>:"/\\|?*]', '_', name)
        return name.strip().strip('.')

    def _reopen_wizard(self):
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_download(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{self._name}/{self.id}/zip_file/{self.zip_filename}?download=true',
            'close': True,
        }
